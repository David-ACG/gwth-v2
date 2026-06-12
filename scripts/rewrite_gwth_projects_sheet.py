"""Rewrite the GWTH Project Ideas spreadsheet into a cleaner v3.

Inputs:  C:\\Users\\david\\Downloads\\All GWTH Project Ideas 21_Apr_26 v2.xlsx
Outputs: C:\\Users\\david\\Downloads\\All GWTH Project Ideas 21_Apr_26 v3.xlsx

Three sheets:
  1. Projects           — restructured master list (capstones + M1/M2/M3 + deprecated)
  2. Lesson-to-Project Map — cross-reference against MONTH_1/2/3_LESSON_IDEAS
  3. Change Log         — what changed and why
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ---------- SHEET 1: PROJECTS ----------
ws = wb.active
ws.title = "Projects"

headers = [
    "ID", "Status", "Month", "Lesson(s)", "Title",
    "Primary Domain", "All Domains", "Short Description",
    "Why / Benefit", "Risks / Learnings", "Primary Tools",
    "Alternative Tools", "Monetisation", "Separate Site?",
    "Complexity", "Dependencies", "Change Notes"
]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="0E7C7B")
section_font = Font(bold=True, color="FFFFFF", size=12)
section_fill_cap = PatternFill("solid", fgColor="4A6CF7")
section_fill_m1 = PatternFill("solid", fgColor="22D3A7")
section_fill_m2 = PatternFill("solid", fgColor="33BBFF")
section_fill_m3 = PatternFill("solid", fgColor="1CBA93")
section_fill_dep = PatternFill("solid", fgColor="9E9A94")
thin = Side(border_style="thin", color="D6DCE6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")


def write_header_row():
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = border


write_header_row()


def section_row(row_idx, label, fill):
    ws.cell(row=row_idx, column=1, value=label).font = section_font
    ws.cell(row=row_idx, column=1).fill = fill
    ws.cell(row=row_idx, column=1).alignment = Alignment(vertical="center")
    for col in range(2, len(headers) + 1):
        ws.cell(row=row_idx, column=col).fill = fill


CAPSTONES = [
    ("Capstone", "1", "L18-L21 (Capstone)",
     "Family AI Bot — Meeting to Tasks/Calendar/Meals/Shopping",
     "familyaibot.com",
     "familyaibot.com, familymanager.net",
     "Record weekly family meeting; transcribe; extract tasks, events, meal plan, shopping list; distribute via email / WhatsApp / calendar. End-to-end M1 pipeline using every primitive.",
     "Solves a universal family pain; teaches end-to-end pipeline without code; flagship M1 asset every student can show.",
     "UK-accent transcription accuracy on noisy audio; calendar-write permissions; family adoption; privacy of recordings.",
     "Whisper (in ChatGPT/Claude) + Claude Cowork + Zapier + Google Calendar + Lovable hub page",
     "ChatGPT Plus + Custom GPT + Zapier free + Google Sheet; OpenClaw for WhatsApp; Fathom for transcription",
     "Free (course asset); hosted SaaS version optional later as lead-gen to GWTH",
     "Yes — consumer-facing site with demo + GWTH CTA",
     "Capstone",
     "L15 agents, L17 automation, L19 transcription, L20 processing",
     "Kept as M1 capstone. Confirmed against MONTH_1_LESSON_IDEAS."),

    ("Capstone", "2", "L16-L19 (Capstone)",
     "AI Readiness Assessment Tool",
     "productarchitect.dev",
     "productarchitect.dev, aibusinesscoach.pro, transformation.diy, productarchitect.co.uk",
     "Production web app: 12-min intake -> deterministic scoring across 6 dimensions -> UK benchmark overlay -> 90-day roadmap -> board-ready PDF -> RAG 'Ask the tool' chat over UK research corpus.",
     "Integrates every M2 skill into one shippable asset; useful on day one (student scores own company); extends into M3 without rewrite; legitimate consulting lead-gen.",
     "Deterministic-core vs LLM-narrative line is the hardest architecture lesson; PDF <-> on-screen parity; RAG citation guardrails; UK data residency (Supabase EU-west-2).",
     "Next.js 16 + Supabase pgvector + Claude Sonnet 4.6 via Vercel AI SDK v5 + @react-pdf/renderer + Sentry + PostHog + Langfuse",
     "Swap LLM via SDK flag (GPT-5, Gemini 3 Pro); Coolify on UK VPS for data-sovereign clients; Vercel free tier default",
     "Lead-gen to GWTH (every assessment recommends enrolment) + optional white-label consulting upsell",
     "Yes — consultancy lead magnet with GWTH upsell",
     "Capstone",
     "L4 APIs, L8 RAG, L9 RAG tuning, L10 Auth+RLS, L11 Agents, L17 RAG chat, L18 PDF/Deploy",
     "**REPLACES** old 'AI Customer-Support Chatbot / askmyco.com' capstone (now demoted to L8 Company Knowledge Bot). Consolidates Row 14 (productarchitect.dev) into the capstone."),

    ("Capstone", "3", "L1-L20 (Capstone — spans month)",
     "AI Readiness Assessment Tool — Enterprise Edition",
     "productarchitect.dev",
     "productarchitect.dev, aibusinesscoach.pro, transformation.diy",
     "Extends the M2 tool: scoring adds McKinsey 12 themes + Sivulka 7 pillars + Rewired 6 capabilities; multi-respondent team mode; Slack + Google Meet invites; CRM lead export; consulting-handover deck; anonymised benchmarking pool; enterprise white-label.",
     "Same codebase as M2 capstone - no rewrite. Students graduate with a genuinely enterprise-grade asset they can sell as consulting. GWTH gets aligned growth channel.",
     "Multi-respondent aggregation UX; CRM rate limits; benchmark opt-in privacy UX; enterprise SSO demands.",
     "M2 stack + HubSpot API + Slack API + Google Meet API + Gamma (or React PDF for slides)",
     "Pipedrive / Salesforce; Notion/Linear for internal tracking; internal-only mode without CRM",
     "Lead-gen + enterprise tier (white-label subscription)",
     "Yes — continuation of productarchitect.dev",
     "Capstone",
     "Complete M2 capstone; M3 L2 12 themes, L6 domains, L14 data advantage, L15 right-to-deploy",
     "**REPLACES** old 'AI Readiness Voice Agent / askevery.one' capstone. askevery.one moved to M2 L13 voice agent build. New capstone evolves from M2 same-codebase."),
]

MONTH_1 = [
    ("Core build", "1", "L11",
     "Save Your Favourite Prompts & Share (Prompt Hub)",
     "cheatprompt.dev",
     "cheatprompt.dev, aicheatprompts.com, aimasterprompt.net, aimasterprompt.store, aipromptsaver.com",
     "Personal prompt library with tags, share links, and copy-to-clipboard. Lovable build with optional auth.",
     "First real Lovable build; teaches prompt discipline; reusable across every other project.",
     "If multi-user, need auth (Supabase) - becomes M2. Copyright on shared prompts.",
     "Lovable",
     "Claude Artifacts (private version); Bolt / v0",
     "Free (personal); Freemium if multi-user (Pro tier = share library) - Pro tier shifts to M2 scope",
     "Yes — consumer site with GWTH CTA",
     "S",
     "L10 First App; L11 Personal Utility",
     "**MERGED** original Row 15 (cheatprompt.dev, M1) + Row 36 (aicheatprompts.com, M2) - both were the same idea."),

    ("Core build", "1", "L16",
     "Work-Wins Log (Brag Manager)",
     "bragmanager.com",
     "bragmanager.com",
     "Custom GPT / Claude Project that logs weekly wins, quantifies them monthly, and drafts promotion-review talking points.",
     "Personal assistant pattern with real annual payoff (pay rise / promotion case); easy ROI to demo.",
     "Privacy - don't feed confidential salary/HR data into public models. Use Claude Project with 'no training' guarantee.",
     "Claude Project (custom instructions) or Custom GPT",
     "Notion AI; Obsidian + Claude",
     "Free (personal tool)",
     "No (personal asset)",
     "S",
     "L2 colleague model; L16 Personal Assistant",
     "Added Primary Tools + monetisation + risk (HR data privacy)."),

    ("Core build", "1", "L16",
     "Personal Goal Planner",
     "goalsfor.me",
     "goalsfor.me",
     "Claude Project that interviews you quarterly on goals, tracks weekly check-ins, nudges when off-track.",
     "Demonstrates long-running context; students see AI hold them accountable over weeks.",
     "Context window - export weekly summaries so the project doesn't bloat.",
     "Claude Project",
     "Custom GPT; Notion AI templates",
     "Free",
     "Optional — GWTH CTA on site",
     "S",
     "L2 colleague model; L16 Personal Assistant",
     "Added tools / monetisation / dependencies."),

    ("Core build", "1", "L14",
     "Token / Credit Usage Monitor (PM5-style)",
     "tokenocd.com",
     "tokenocd.com, trackaicredits.com, creditlogger.com",
     "Lovable dashboard that reads usage from ChatGPT / Claude / Gemini subscription pages and visualises burn-rate like a Concept2 PM5 rower. Includes a pacer.",
     "Teaches dashboard-building; solves a real pain (am I about to hit my cap?); highly shareable / viral.",
     "Scraping subscription dashboards may need a browser-use agent (M2). M1 version uses manual CSV paste.",
     "Lovable (M1 version — manual paste) / Claude Artifacts for the gauge",
     "v0; Bolt; Cursor + Next.js (M2 version)",
     "Free (M1) -> Freemium when scraping added (M2)",
     "Yes — community tool + GWTH CTA",
     "S (M1) / M (M2 live-scrape)",
     "L10 First App; L11 Personal Utility; L14 Dashboard",
     "Kept M1. Marked M2 extension for live-scrape version."),

    ("Core build", "1", "L12",
     "My Website (personal brand / portfolio)",
     "eyeonai.dev",
     "eyeonai.dev, gcsedad.com",
     "Personal blog / landing page built in one hour with Lovable. Students deploy to a real domain (£5-10 UK registrar).",
     "Every student leaves M1 with a live URL - forced demonstration of capability.",
     "Domain DNS confusion; HTTPS setup; content ownership.",
     "Lovable + Vercel / Cloudflare Pages + UK registrar (Krystal, 123-reg, Fasthosts)",
     "Bolt; v0; WordPress for blog-heavy",
     "Free (personal)",
     "Yes — personal brand",
     "S",
     "L12 My Website",
     "Consolidated generic 'blog site' rows into one L12 build template."),

    ("Core build", "1", "L6 + L12",
     "Research + Comparison Site Template",
     "bestrobotmop.com",
     "bestrobotmop.com, bestelectricblankets.com, bestgcseapps.com, comparemultivitamins.com, cooksgadgets.com, dofechecklist.com, dogwalksnear.me, passthewine.com, barkwelcome.com, bestdogfriendlyholidays.com, bestdogfriendlypubs.com, bestmtd.com, comparemtd.co.uk, topmtd.co.uk",
     "AI-researched, AI-written comparison / niche research site on a specific topic. Lovable or v0 build; content drafted in Claude and pasted in. Affiliate-link ready.",
     "Template that produces 10+ niche sites - real revenue potential via affiliates; teaches research (L6) + website (L12) stacked.",
     "**Affiliate disclosure** required under UK CAP Code. Do not fabricate product claims - Claude tends to hallucinate model numbers.",
     "Lovable / v0 + Claude research + UK registrar",
     "Bolt; WordPress with AI plugin",
     "Paid (affiliate commission) - each site earns independently; GWTH CTA on each",
     "Yes — each niche gets its own subdomain or site",
     "S per site, done in one lesson",
     "L6 research; L12 website",
     "**Consolidated** 14 comparison-site rows (Rows 17, 33, 60-69) into ONE reusable template with domain list. Added affiliate / CAP Code risk."),

    ("Core build", "1", "L6",
     "Research -> Executive Deck Generator",
     "firstcalldeck.com",
     "firstcalldeck.com",
     "Claude researches a topic, drafts a 10-slide deck brief; student exports to Gamma / Beautiful.ai / Tome; resulting deck shipped in under an hour.",
     "Sales / partnership pre-meeting prep compressed from 4 hours to 40 minutes.",
     "Copyright on sourced imagery; UK IPO text-and-data-mining rules still in flux.",
     "Claude + Gamma",
     "ChatGPT + Beautiful.ai; Tome",
     "Freemium (free template, Pro tier = custom branding)",
     "Yes — GWTH CTA",
     "S",
     "L6 Research; L7 Content; L8 Visuals",
     "Added tools, dependency mapping, CAP Code risk."),

    ("Core build", "1", "L15 + L17",
     "Grocery Shopping Agent (Family AI Bot extension)",
     "groceryshoppingagent.com",
     "groceryshoppingagent.com, smrtlist.com, shopgrocery.co.uk, smrtlst.com, supermarketshopping.com",
     "Zapier / Claude Cowork agent that takes the shopping list from the Family AI Bot capstone and pushes into Tesco / Ocado / Sainsbury's basket. Optional: voice confirmation via Siri Shortcut.",
     "Genuine time-saver; demonstrates the 'agent' primitive at M1 level without code.",
     "Retailer T&Cs - some UK grocers block automated basket-building. Use Claude for Chrome as polite fallback.",
     "Claude Cowork / Zapier + Google Sheet + UK grocer web (via Claude for Chrome)",
     "Make.com Maia; OpenClaw in M2",
     "Free (extension of capstone)",
     "Yes — extension of familyaibot",
     "M",
     "L15 Agents; L17 Automation; L20 Capstone",
     "**Merged** Row 24 (grocery agent) + Row 55 (smrtlist.com et al domains). All are the same idea."),

    ("Core build", "1", "L15 + L17",
     "Recipe -> Shopping List Agent (Family AI Bot extension)",
     "recipeunboxed.com",
     "recipeunboxed.com",
     "Given a set of recipes, agent extracts ingredients, de-duplicates across the week, and feeds the grocery agent.",
     "Completes the meal-plan loop students start in the capstone.",
     "OCR on recipe photos - quality varies.",
     "Claude + Zapier + Google Sheet",
     "Cowork; Make.com Maia",
     "Free",
     "Yes — familyaibot extension",
     "S",
     "L17 Automation; L19 Extraction; L20 Capstone",
     "Kept M1, linked to capstone."),

    ("Core build", "1", "L12",
     "Party Photo Sharing",
     "partysnapper.com",
     "partysnapper.com",
     "Simple Lovable site where party guests upload photos to a shared gallery; host gets a zipped download and a highlight reel draft.",
     "Demonstrates file uploads + AI image summary; real family / wedding use case.",
     "GDPR on uploaded faces; need consent banner.",
     "Lovable + Cloudflare R2 / Supabase Storage (M1 manual link)",
     "Bolt; v0",
     "Freemium (free = 10 photos, paid = unlimited)",
     "Yes",
     "S",
     "L12 Website",
     "Added GDPR / consent risk."),

    ("Core build — REBUILD", "1", "L6 + L12",
     "AI Course Comparison Directory",
     "aicoursedirectory.com",
     "aicoursedirectory.com, aicoursedriectory.com (typo variant), bestcoursetools.com, bestcoursebuilder.com",
     "Lovable site comparing AI course offerings (Coursera, Maven, DeepLearning.AI, GWTH, GoCoach, Section). Includes pricing, duration, level, outcomes, review aggregator.",
     "SEO-friendly; natural GWTH channel; low-maintenance. David already owns the domain — rebuild from scratch, not compete.",
     "**David's existing site went stale** — restart from zero. Affiliate disclosure (CAP Code). Keep content fresh via weekly research agent (M2 extension).",
     "Lovable + Claude research",
     "v0; WordPress; Astro",
     "Paid (affiliate) / Lead-gen for GWTH",
     "Yes",
     "S",
     "L6 Research; L12 Website",
     "Promoted Row 51 to M1 core build. Per David: existing site stale, rebuild from scratch using aicoursedirectory.com."),

    ("Optional / personal", "1", "L6 + L12",
     "Italian Citizenship by Descent Guide",
     "(David to choose)",
     "(David to choose)",
     "Research + blog / guide site on iure sanguinis Italian citizenship for UK residents.",
     "Personal project; demonstrates niche research at M1 level.",
     "Legal content - must mark 'not legal advice' and cite Italian consulates.",
     "Lovable + Claude research",
     "v0; WordPress",
     "Freemium (free guide, paid = checklist download)",
     "Yes",
     "S",
     "L6 Research; L12 Website",
     "Retained from Row 53; mark as Optional."),
]

MONTH_2 = [
    ("Core build", "2", "L8 + L9",
     "AI Customer-Support Chatbot (Company Knowledge Bot)",
     "askmyco.com",
     "askmyco.com, notesrag.com",
     "RAG chatbot trained on company docs. Build the pgvector index, ingest PDFs / website, ground all answers in citations, embed into client website.",
     "Highest-value consulting offer for UK SMEs (~£5-20k engagements); repeatable across clients.",
     "Citation guardrail must be rock-solid (never fabricate sources); UK data residency via Supabase EU-west-2; GDPR DPA with client.",
     "Next.js 16 + Supabase pgvector + Claude Sonnet 4.6 via Vercel AI SDK v5 + Cohere Rerank v3",
     "GPT-5 / Gemini 3 Pro via same SDK; LangChain for advanced pipelines; Qdrant if UK-VPS-hosted",
     "Paid (SaaS or consulting deliverable); Lead-gen when used on own site",
     "Yes — demo site + GWTH CTA",
     "L",
     "L6 ingest; L7 embeddings; L8 RAG; L9 tuning",
     "**DEMOTED** from M2 Capstone to L8 Company Knowledge Bot - replaced as capstone by AI Readiness Assessment Tool."),

    ("Core build", "2", "L8 + L9",
     "Transcripts RAG (Company Meeting Knowledge)",
     "aitranscriptionhub.com",
     "aitranscriptionhub.com",
     "Ingest transcripts of meetings / podcasts / YouTube; RAG over them; ask questions like 'what did Alice say about pricing last quarter?'.",
     "Every team has a pile of Otter / Fathom transcripts going unused.",
     "Transcript accuracy on UK accents; speaker diarisation; GDPR on captured voices.",
     "Docling / unstructured + Supabase pgvector + Claude",
     "LlamaIndex; Qdrant",
     "Paid SaaS (£50-200/mo per team) - exactly the sort of tool where token cost must be covered",
     "Yes — SaaS separate from GWTH",
     "M",
     "L6 ingest; L8 RAG; L10 Auth+RLS",
     "**MOVED** from M3 to M2 L8. Was mismarked as Month 3."),

    ("Core build", "2", "L8 + L9",
     "Medical RAG — c.diff Research Guide",
     "cdiffguide.com",
     "cdiffguide.com",
     "RAG over NHS, CDC and peer-reviewed c.diff research; patients / carers ask plain-English questions; every answer cites NHS / published source.",
     "High personal / family relevance; demonstrates responsibility with medical content.",
     "**MHRA / NHS rules** - mark clearly 'not medical advice'; never suggest treatment; refuse dosing questions. Consider DPA if collecting user questions.",
     "Next.js + Supabase pgvector + Claude Sonnet 4.6 + Cohere Rerank",
     "GPT-5; Gemini 3 Pro; Qdrant",
     "Free (public service) with GWTH CTA - do NOT paywall medical info",
     "Yes — public service site",
     "M",
     "L5 Security/GDPR; L8 RAG; L9 tuning",
     "Added MHRA risk + free monetisation."),

    ("Core build", "2", "L8 + L9",
     "Employment Law RAG (UK)",
     "employmentlawbuddy.com",
     "employmentlawbuddy.com",
     "RAG over gov.uk ACAS guidance, Employment Rights Act 1996, tribunal decisions. Workers ask plain-English questions; every answer cites statute or case.",
     "Highest ROI for a UK SME that can't afford a retained lawyer.",
     "**Not legal advice** disclaimer mandatory; SRA rules on reserved legal activities; do NOT draft contracts for clients (reserved activity).",
     "Next.js + Supabase pgvector + Claude + Cohere Rerank",
     "GPT-5; Gemini 3 Pro; LlamaIndex",
     "Freemium - Free for individuals; Pro for HR/manager tier (£15-30/mo)",
     "Yes — SaaS + GWTH CTA",
     "M",
     "L5 Security/GDPR; L8 RAG; L9 tuning; L10 Auth",
     "Added SRA risk + freemium monetisation."),

    ("Core build", "2", "L10",
     "Health Logging App (GDPR special-category data)",
     "healthlog.app",
     "healthlog.app",
     "Multi-user app where users log sleep, mood, symptoms, meds; Supabase RLS enforces per-user isolation; AI weekly summary.",
     "Demonstrates GDPR-special-category handling - a hard-fail pattern students must learn.",
     "**Special-category health data** under UK GDPR - need explicit consent, DPIA, and DPA with hosting provider. RLS must be bulletproof.",
     "Next.js + Supabase Auth + RLS + pgvector (for summary search)",
     "Better Auth; Clerk (US-based — watch data residency)",
     "Freemium - free tier with cap, Pro = unlimited + export",
     "Yes",
     "M",
     "L5 Security; L10 Auth+RLS",
     "Added special-category GDPR risk."),

    ("Core build", "2", "L14",
     "Physical Whiteboard -> Online Meeting Share",
     "whiteboard.bot",
     "whiteboard.bot, hybridwb.com, whiteboard.show",
     "Phone camera streams physical whiteboard; Claude multimodal cleans, OCRs, and posts the structured text to Teams/Zoom chat every 30s.",
     "Hybrid-meeting pain; no incumbent tool solves it well.",
     "Camera permissions; low-light OCR; handwriting accuracy.",
     "Claude 4.6 multimodal + WebRTC + Supabase Realtime",
     "GPT-5 Vision; Gemini 3 Pro; open-source OCR (Tesseract fallback)",
     "Paid SaaS (team seat) - token cost is material",
     "Yes",
     "L",
     "L14 Multimodal; L10 Auth",
     "Kept M2."),

    ("Core build", "2", "L12",
     "Share Trajectory (Research, NOT Advice)",
     "sharetrajectory.com",
     "sharetrajectory.com, sharepies.com, stocktrajectory.com",
     "Research agent that gathers news, filings, analyst notes on a ticker; produces a research pack. **Positioning: research only, not financial advice.**",
     "UK retail investor research compression; free alternative to Bloomberg.",
     "**FCA regulation** - must NOT give advice / recommendations. Disclaimer + no 'buy/sell' language. Check financial-promotion rules if shared publicly.",
     "Claude for Chrome (research agent) + Next.js",
     "Perplexity API; Gemini 3 Pro Deep Research",
     "Freemium - free research, paid = PDF export + history",
     "Yes",
     "M",
     "L12 Research Agent; L5 Regulatory awareness",
     "Added FCA risk + research-not-advice positioning."),

    ("Core build", "2", "L12",
     "Stag / Hen Guide (Travel Planning Agent)",
     "stagsguide.co.uk",
     "stagsguide.co.uk, hensguide.co.uk",
     "Voice-led planner: group leader describes the group + budget + dates; agent proposes itinerary (venues, transport, accommodation) with UK-specific options (Edinburgh, Brighton, Bristol).",
     "Commercial vertical; seasonal demand; affordable CPC.",
     "Hotel / bar T&Cs block automated booking - present options, let user book manually.",
     "Claude for Chrome + Next.js + Supabase",
     "Browserbase; Stagehand",
     "Freemium or Paid (one-off £20-50 per group)",
     "Yes",
     "M",
     "L12 Research Agent; L10 Auth",
     "Kept M2."),

    ("Core build", "2", "L13 + L14 + L15",
     "Video CV Builder",
     "myvideo.cv",
     "myvideo.cv, vividcv.com, vividprofile.com, govideo.cv",
     "User uploads CV; AI generates script; ElevenLabs voice or user reads; AI interview segments; ffmpeg assembles; Stripe paywall; analytics on views.",
     "High-value graduate / job-seeker product; ElevenLabs + Stripe + analytics all in one build.",
     "Deepfake risk - require real-face verification before paying. GDPR on uploaded CV content.",
     "Next.js + Supabase + ElevenLabs Agents + ffmpeg + Stripe",
     "OpenAI Voice Agents; Vapi; Mux for video",
     "Paid - £20 one-off or £50/yr subscription. Token cost is material.",
     "Yes",
     "L",
     "L10 Auth; L13 Voice; L14 Multimodal; L15 Stripe/Supabase",
     "Kept M2; added deepfake + GDPR risk + Stripe dependency."),

    ("Core build", "2", "L13",
     "Spanish / Italian GCSE Voice Tutor",
     "spanish.gcseapp.com",
     "gcseapp.com (parent), spanish.gcseapp.com, italian.gcseapp.com (future), bestgcseapps.com, gcsequestions.com, revisionamigo.com",
     "Conversational voice app that role-plays a Pearson Edexcel GCSE Spanish examiner. Asks curriculum-aligned questions; marks vocabulary and fluency. Italian variant for holiday prep.",
     "Solves a real family pain (David's daughter); ElevenLabs or OpenAI Realtime with 300k credits David already has. Scales across languages as gcseapp.com subdomains.",
     "LLM Spanish fluency varies - test with native speaker; examiner tone must be friendly not intimidating.",
     "ElevenLabs Agents (300k credits) + Claude Max (existing subscription) + Next.js PWA",
     "OpenAI Realtime (Voice mode); Vapi; Gemini 2.5 Pro Voice",
     "Free (personal) - could become Paid (£5-10/mo) as tutor SaaS. Mark tokens carefully: voice is expensive.",
     "Yes — subdomain architecture: spanish.gcseapp.com, italian.gcseapp.com, etc.",
     "M",
     "L4 APIs; L13 Voice Agents",
     "Prototype already built per David's note. Domain assigned per David 21/4: spanish.gcseapp.com. Parent gcseapp.com gives clean language-subdomain pattern."),

    ("Core build", "2", "L13",
     "AI Readiness Voice Agent (Interview Intake)",
     "askevery.one",
     "askevery.one, voicescope.io",
     "Voice-led intake for the AI Readiness Assessment Tool: leader chats with a voice agent for 8 mins instead of filling a form. Transcript feeds the scoring engine.",
     "Higher completion rate vs form; enterprise leaders prefer talking.",
     "Voice minute cost (~£0.10-0.25/min); PII handling; UK voice model availability.",
     "ElevenLabs Agents + Vapi + Next.js + Supabase",
     "OpenAI Realtime; Retell; Pipecat",
     "Bundled with Capstone - not separately monetised",
     "No — feature of productarchitect.dev",
     "M",
     "L4 APIs; L10 Auth; L13 Voice; Capstone",
     "**MOVED** from M3 Capstone to M2 L13 voice agent build. Now a feature that feeds the Assessment Tool."),

    ("Optional / Lab", "2", "L13 + L14",
     "AI-Generated Sleep Music & Distribution",
     "talktosleep.com",
     "talktosleep.com, soporific.app",
     "Generate ambient sleep tracks (Suno) + album art (GPT Image 1.5 / FLUX 2) + promo site (Lovable) + distribute to Spotify via DistroKid. Optional: AI-narrated sleep stories (ElevenLabs).",
     "Tangible shippable product; passive income potential; demonstrates Content Creation primitive.",
     "Suno / music-gen tools NOT in prescriptive big-3 stack - mark as Optional/Lab. Streaming royalties small; copyright of AI music unclear globally and in UK.",
     "Suno + Lovable + DistroKid + ElevenLabs",
     "Udio; AIVA; SoundRaw",
     "Paid (streaming royalties, DistroKid £23/yr); marginal",
     "Yes",
     "M",
     "L14 Multimodal (audio generation)",
     "**RECLASSIFIED** as M2 Optional / Lab - Suno is not in the prescriptive big-3 stack."),

    ("Core build", "2", "L14 + L15",
     "SEO for AI Bots (Generative Engine Optimisation)",
     "sitegeo.net",
     "sitegeo.net, pixelseo.co.uk, seoappsdirectory.com",
     "Audit a site's visibility inside ChatGPT / Claude / Perplexity / Gemini answers; produce fixes (schema.org markup, entity clarity, citation-worthy content); weekly visibility report.",
     "New category - 'GEO' is a 2026 topic; SMEs budget-shifted ~£400/mo from SEO agencies to GEO.",
     "Category is <12 months old - positioning drifts; LLM providers change ranking signals frequently.",
     "Next.js + Claude for Chrome (tests how each LLM answers queries) + n8n schedule",
     "Perplexity API; OpenAI browse mode; Gemini Deep Research",
     "Paid SaaS (£100-400/mo per site); Lead-gen to GWTH consulting",
     "Yes",
     "L",
     "L12 Research Agent; L15 Automation",
     "Kept M2."),

    ("Core build", "2", "L10 + L13",
     "Job Try-Out (Interview + Skills Test Automation)",
     "jobtryout.co.uk",
     "jobtryout.co.uk, jobtryouts.co.uk",
     "Candidate takes a 20-min voice interview + a tech skills test against a rubric the company set. Company sees scored results + transcript. Removes risk both sides.",
     "Real HR vertical; SaaS pricing; UK employment-law-friendly (both sides consent up-front).",
     "**UK Employment Law** - bias audit required; must not be the sole decision input; DSAR on candidate data.",
     "Next.js + Supabase + ElevenLabs Agents + Claude scoring",
     "Vapi; Retell; OpenAI Realtime",
     "Paid SaaS (per-interview fee £10-25 + company subscription)",
     "Yes",
     "L",
     "L5 GDPR / UK employment law; L10 Auth; L13 Voice; L11 Agent-scoring",
     "Added Employment-Law + bias-audit risk."),

    ("Core build", "2", "L20",
     "Fractional Exec / Buddy Platform",
     "fractionalbuddy.com",
     "fractionalbuddy.com, fractional.express, fractional.top, fractionalhub.co.uk",
     "Directory + matching for fractional CTOs/CFOs/CMOs in the UK. Pair AI-led onboarding (collects company profile) + human introductions + RAG-assist during engagement.",
     "Prototype already built; customers interested; aligns with L20 Consulting Service Page.",
     "IR35 / umbrella company rules for UK fractionals; insurance (PI); introducer-fee VAT.",
     "Next.js + Supabase + Claude + Stripe",
     "Lovable for MVP; Webflow",
     "Paid (introducer fee or subscription)",
     "Yes",
     "L",
     "L8 RAG; L10 Auth; L15 Stripe; L20 Consulting Service",
     "Kept M2. Prototype flagged. Added IR35 / PI risk."),

    ("Core build", "2", "L7 + L11 + L15",
     "Automated AI News Scraper + Daily Newsletter",
     "newsbot.biz",
     "newsbot.biz, latestgreatest.news, eyeonai.dev",
     "Agent scrapes AI news sources 2x/day; dedups; LLM summarises; ranks by interest; sends top-3 newsletter. **Drop the Digg-style voting** - David's earlier test showed zero incentive to vote.",
     "Curate-not-create content engine; drives traffic to GWTH.",
     "**DROP voting feature** - tested by David, didn't work. Editorial tone curation; copyright of summarised sources (fair-dealing UK).",
     "n8n 2.0 + Claude + Supabase + SendGrid / Resend",
     "Make.com Maia; Zapier AI Agents",
     "Free (sponsored-by-GWTH model) - Lead-gen",
     "Yes",
     "M",
     "L7 Embeddings (dedup); L11 Agents; L15 Automation",
     "**MOVED** from M1 to M2. Added David's learning (voting didn't work -> drop it)."),

    ("Core build", "2", "L7 + L11 + L15",
     "Blog Posts from YouTube + Tweets",
     "blogbot.app",
     "blogbot.app, blogbuddy.co",
     "Agent watches a list of YouTube channels + X accounts; transcribes new videos; drafts blog posts; posts to your CMS. n8n or Claude Skills under the hood.",
     "Real marketing ROI; reusable pattern; teaches the full agent-automation loop.",
     "Copyright of transcribed content (fair-dealing UK); attribution; risk of AI slop if unchecked.",
     "n8n 2.0 + Claude Skills + YouTube API + Twitter/X API + WordPress / Ghost API",
     "Make.com; Zapier AI Agents",
     "Freemium (self-host free; managed = paid)",
     "Yes",
     "M",
     "L7 Embeddings; L11 Agents; L15 Automation",
     "**MOVED** from M1 to M2 - requires agents + MCP / Claude Skills."),

    ("Core build", "2", "L11",
     "Offline Claude Code Sync Tool (Train Coder)",
     "traincoder.com",
     "traincoder.com",
     "Desktop utility: sync last Claude Code reply per tab for offline (train travel); compose next 5 prompts across 5 tabs; auto-replay on reconnect. Built on syncthing.",
     "Solves David's own UK-commuter pain; niche but real audience.",
     "Claude Code / Cursor internal APIs are not public - reverse-engineer carefully, may break on updates. Warp vs Ghostty vs iTerm choice.",
     "Tauri / Electron + syncthing + localwhisper (for STT) or AnythingLLM",
     "VS Code extension; Cursor fork",
     "Free (OSS) - student productivity tool",
     "Yes — GWTH CTA",
     "L",
     "L11 Agents; L15 Automation",
     "Domain assigned per David 21/4: traincoder.com."),

    ("Lab / Optional", "2", "Lab",
     "Private Superwhisper Clone (local STT)",
     "lowhisper.com",
     "lowhisper.com, localwhisper.dev, mywhisper.dev",
     "Run Whisper locally (Mac/Linux) for private voice-to-text; GlobalHotkey paste to any app; offline.",
     "Data-sovereign alt to SuperWhisper / Whisprflow; teaches local-model deployment.",
     "Not in prescriptive M2 core stack - Lab only. Mac Silicon-accelerated; Windows has PyTorch path.",
     "whisper.cpp + Python launcher + OS hotkey",
     "WhisperX; Faster-Whisper; Distil-Whisper",
     "Free (OSS)",
     "Yes",
     "M",
     "L2 Dev Env; L11 Agents",
     "**MOVED** from M1 to M2 Lab - requires local-model tooling not in M1."),

    ("Core build", "2", "L4",
     "LLM API Playground & Cost Comparison (aiapi.click)",
     "aiapi.click",
     "aiapi.click, apicompare.net, llmapi.click, llmapicompare.com",
     "Public playground: paste one prompt, pick model (Claude Sonnet 4.6, Opus 4.7, GPT-5, Gemini 3 Pro, Haiku 4.5), get side-by-side streamed outputs with live token count and £ cost per call. Saved prompts; share-URL per comparison.",
     "Teaches Vercel AI SDK v5 provider-swap + streaming + cost tracking in one sitting. Student ships a genuinely useful public tool — comparison sites rank well in LLMs. Aligns with the 'Big-3 + challenger' GWTH stack rule.",
     "API key rotation (never expose client-side); rate-limit handling across providers (Claude 50k TPM, GPT-5 higher); cost column must match provider billing.",
     "Next.js 16 + Vercel AI SDK v5 + Anthropic / OpenAI / Google SDKs + Vercel Hobby",
     "Direct REST for each provider; Langfuse for trace recording",
     "Freemium — free tier caps per-session spend; Pro adds private workspaces",
     "Yes — public utility with GWTH CTA",
     "S",
     "L2 Dev Env; L4 APIs",
     "**REFRAMED** per David 21/4 ('too generic'): now named project 'LLM API Playground' on aiapi.click. Reuses 4 of David's existing domains. Still teaches L4 API-calling primitives."),

    ("Core build — Phase 1 of askmyco.com", "2", "L6 (Phase 1)",
     "Document Ingestion Pipeline — Phase 1 of askmyco.com",
     "askmyco.com",
     "askmyco.com, notesrag.com",
     "Phase 1 of the Company Knowledge Bot: upload client PDF / Word / Excel / website URLs; produce clean Markdown + JSON chunks ready for embedding. Wizard UI + preview. Same codebase carries forward to L8 (RAG) and L9 (tuning).",
     "Students see the full RAG lifecycle on ONE named project, not three abstract demos. Document ingestion is where 80% of RAG-in-production fails — worth a dedicated lesson.",
     "Scanned PDFs need OCR (Tesseract fallback); tables often mangled; DOCX metadata variance; UK data residency (Supabase EU-west-2 from day one).",
     "Docling + Next.js + Supabase Storage + Supabase pgvector (prep)",
     "Unstructured.io; LlamaParse; textract",
     "Bundled with askmyco.com (no separate monetisation)",
     "No — phase of askmyco.com",
     "S",
     "L2 Dev Env; L6 Ingest",
     "**REFRAMED** per David 21/4 ('too generic'): now Phase 1 of askmyco.com. L6 ingests -> L8 RAG -> L9 tunes, one project across three lessons."),

    ("Core build — Phase 1 of fractionalbuddy.com", "2", "L11 (Phase 1)",
     "Fractional Inbox Assistant — Phase 1 of fractionalbuddy.com",
     "fractionalbuddy.com",
     "fractionalbuddy.com, fractional.express, fractional.top, fractionalhub.co.uk, fractionalned.com, fractionalroles.co.uk",
     "Phase 1 of the Fractional platform: Claude Agent SDK + MCP Gmail connector triages a fractional exec's multi-client inbox, classifies by client + urgency, drafts replies per client tone, runs daily 7am. Teaches CLIs-first + MCP-second pattern against a real pain point (fractionals juggle 4-8 client inboxes).",
     "Showcases Anthropic Agent SDK + MCP ecosystem on a genuine persona (UK fractional CTO/CFO market growing 2026). Phase 2 later adds the marketplace + Stripe.",
     "OAuth permissions (per-client tokens); draft-only mode until trust established; multi-tenant isolation.",
     "Claude Agent SDK + MCP Gmail + gh CLI + Cloudflare Worker (schedule) + Supabase (client profiles)",
     "n8n with Gmail trigger; Make.com AI Agents; Zapier AI Agents",
     "Bundled with fractionalbuddy.com (Paid once platform layer added in L20)",
     "No — phase of fractionalbuddy.com",
     "M",
     "L2 Dev Env; L4 APIs; L5 Security; L11 Agents",
     "**REFRAMED** per David 21/4 ('too generic'): now Phase 1 of fractionalbuddy.com. Real persona (fractional exec) replaces abstract demo."),

    ("Core build", "2", "L12",
     "UK Planning Applications Research Agent",
     "planningapplicationbot.co.uk",
     "planningapplicationbot.co.uk",
     "Claude for Chrome agent scrapes local UK council planning portals; notifies user when new applications affect their postcode/property.",
     "UK-specific value; real pain for architects + property owners.",
     "Council portals vary massively; scraping T&Cs; re-sale of public data (mostly OK, check licensing).",
     "Claude for Chrome + Next.js + Supabase + scheduled Cloudflare Worker",
     "Browserbase + Stagehand; Apify; ScrapingBee",
     "Freemium (free monitoring, Paid = notifications + history)",
     "Yes",
     "M",
     "L11 Agents; L12 Browser Agents",
     "Domain assigned per David 21/4: planningapplicationbot.co.uk."),

    ("Core build — Phase 3 of sitegeo.net", "2", "L15 (Phase 3)",
     "Production Automation — Phase 3 of sitegeo.net",
     "sitegeo.net",
     "sitegeo.net, pixelseo.co.uk, seoappsdirectory.com",
     "Phase 3 of the GEO SaaS. Self-hosted n8n 2.0 workflow: Stripe webhook -> Supabase subscription update -> Claude welcome email -> Slack alert on new sign-ups -> weekly visibility-report automation. Takes the L14 GEO audit engine and productionises it as a paid £100-400/mo SaaS.",
     "Every M2 SaaS build reuses this automation pattern, but teaching it against sitegeo.net makes the billing / subscription loop feel real, not abstract. Students ship a genuinely paid SaaS at end of Week 7.",
     "n8n 2.0 hosting (UK Hetzner / Coolify); Stripe webhook signing; retry-on-failure semantics; VAT MOSS for UK/EU customers.",
     "n8n 2.0 (self-hosted Coolify) + Stripe + Supabase + Resend/SendGrid",
     "Make.com; Zapier Enterprise; Autumn (Stripe wrapper)",
     "Paid — this is the lesson where sitegeo.net goes from demo to revenue",
     "No — phase of sitegeo.net",
     "M",
     "L2 Dev Env; L10 Auth; L14 GEO engine; L15 Automation",
     "**REFRAMED** per David 21/4 ('too generic'): now Phase 3 of sitegeo.net. L14 builds the GEO engine; L15 wires it into Stripe + n8n + Supabase so it ships as a paid SaaS."),
]

MONTH_3 = [
    ("Core build", "3", "L10 + M3 Mini-Build 4",
     "Internal AI Skills Library (Dojo Clone)",
     "aiconfident.io",
     "aiconfident.io (primary), aiserviceindex.com, corporateaihub.com, vibeappmarket.com (alt)",
     "Internal marketplace where team members publish reusable AI skills (prompts / workflows / Claude Code skill files / MCP configs). Sensei LLM recommends top 5 for each user's role.",
     "Operationalises 'Harness Engineering' at SME scale; answer to 'where are all our good prompts?'. aiconfident.io positions it as a confidence-building tool for enterprises starting their AI journey.",
     "Quality curation (don't let the library become full of broken prompts); search relevance on small corpora.",
     "Next.js + FastAPI + Qdrant (or Supabase pgvector) + Anthropic SDK",
     "Claude Projects share; Notion database (low-tech version)",
     "Paid (per-seat team SaaS) - Lead-gen to GWTH enterprise",
     "Yes — enterprise-facing + GWTH CTA",
     "L",
     "M3 L10 Ramp Glass / Harness Engineering",
     "Domain assigned per David 21/4: aiconfident.io primary (with aiserviceindex.com and corporateaihub.com as alternatives)."),

    ("Core build", "3", "L5 + L6 + M3 Mini-Build 2",
     "Domain Discovery Assistant (AI Economic Leverage)",
     "ideagrading.com",
     "ideagrading.com, primeusecase.com",
     "Conversational LLM interviews exec about business (10 Q's); returns prioritised domain map: top 3-5 places where AI would cascade economically.",
     "Replaces the 2-hour 'which use case?' workshop SMEs waste time on.",
     "LLM must handle sector-specific leverage points (fraud for banks, pricing for airlines, supply chain for manufacturers).",
     "Next.js + Vercel AI SDK (streaming) + Claude Sonnet 4.6 (JSON structured output)",
     "GPT-5; Gemini 3 Pro",
     "Bundled with Capstone (feature of productarchitect.dev) OR standalone Freemium",
     "Optional — feature of Capstone",
     "M",
     "M2 Capstone tech; M3 L5 Leverage; L6 Domains",
     "Kept M3. Matches Mini-Build 2."),

    ("Core build", "3", "L14 + L15 + M3 Mini-Build 3",
     "Right-to-Deploy Generator (Governance Artefacts)",
     "governancekit.co.uk",
     "governancekit.co.uk",
     "Form -> generates model card, risk register, fairness review, monitoring plan, kill switch procedure. UK (FCA/ICO) / DBS PURE / McKinsey Rewired framings.",
     "Highest ROI enterprise deliverable - compresses compliance days into minutes in regulated UK sectors. governancekit.co.uk positions it as a UK-first compliance toolkit.",
     "Regulatory output must match actual statute - cite five FCA/ICO principles by name; sectors expect specific wording.",
     "Next.js + FastAPI + Supabase + @react-pdf/renderer + non-LLM reviewer pass",
     "LangChain; OpenAI structured outputs",
     "Paid (enterprise tier) - Lead-gen",
     "Yes — enterprise SaaS",
     "L",
     "M2 L5 GDPR/ICO; M3 L14 Data; L15 Right to Deploy",
     "Domain assigned per David 21/4: governancekit.co.uk."),
]

DEPRECATED = [
    ("Duplicate — merged", "—", "—",
     "Hub for AI Prompts (aicheatprompts.com)",
     "aicheatprompts.com",
     "aicheatprompts.com",
     "(Merged into M1 cheatprompt.dev - same idea.)",
     "", "", "", "", "", "", "", "",
     "**DUPLICATE** of Row 15 (cheatprompt.dev). Merged - aicheatprompts.com added to that row's All Domains."),

    ("Placeholder — needs decision", "—", "—",
     "(unnamed) teamdiary.com",
     "", "teamdiary.com", "", "", "", "", "", "", "", "", "",
     "Row 56 - no title or description. Could fit M3 Internal AI Skills Library. David to allocate or drop."),

    ("Placeholder — needs decision", "—", "—",
     "(unnamed) techfromchina.com",
     "", "techfromchina.com", "", "", "", "", "", "", "", "", "",
     "Row 57 - no title or description. David to allocate or drop."),

    ("Placeholder — needs decision", "—", "—",
     "(unnamed) vibeappmarket.com",
     "", "vibeappmarket.com", "", "", "", "", "", "", "", "", "",
     "Row 58 - now listed as secondary/alt domain for M3 Dojo (aiconfident.io primary per David 21/4). Could also host 'best vibe-coding apps' marketplace using domains bestvibecoding.tools, bestvibecodingapp.com."),
]


current_row = 2

section_row(current_row, "CAPSTONES", section_fill_cap)
current_row += 1

id_counter = 1


def write_project(tpl):
    global current_row, id_counter
    row = [id_counter] + list(tpl)
    for col_idx, val in enumerate(row, start=1):
        c = ws.cell(row=current_row, column=col_idx, value=val)
        c.alignment = wrap
        c.border = border
        c.font = Font(size=10)
    current_row += 1
    id_counter += 1


for p in CAPSTONES:
    write_project(p)

section_row(current_row, "MONTH 1 — AI FOR YOUR LIFE (small-to-medium builds)", section_fill_m1)
current_row += 1
for p in MONTH_1:
    write_project(p)

section_row(current_row, "MONTH 2 — AI FOR YOUR INDUSTRY (most builds; production-grade)", section_fill_m2)
current_row += 1
for p in MONTH_2:
    write_project(p)

section_row(current_row, "MONTH 3 — AI FOR YOUR COMPANY (strategy + extension builds)", section_fill_m3)
current_row += 1
for p in MONTH_3:
    write_project(p)

section_row(current_row, "DEPRECATED / PLACEHOLDERS (needs David's decision)", section_fill_dep)
current_row += 1
for p in DEPRECATED:
    write_project(p)

widths = [4, 20, 7, 22, 40, 22, 45, 55, 45, 40, 35, 32, 28, 22, 10, 32, 55]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.freeze_panes = "A2"


# ---------- SHEET 2: LESSON MAP ----------
ws2 = wb.create_sheet("Lesson-to-Project Map")
ws2_headers = ["Month", "Lesson", "Lesson Title", "Build (from lesson ideas doc)", "Project(s) in spreadsheet", "Status"]
for col_idx, h in enumerate(ws2_headers, start=1):
    c = ws2.cell(row=1, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = border

M1_LESSONS = [
    ("1", "L3", "Your AI Toolkit", "Tool Log (Markdown)", "—", "Generic — no domain project"),
    ("1", "L7", "Content Creation — text", "Content Sprint", "—", "Generic — no domain project"),
    ("1", "L8", "Content Creation — visuals/audio/video", "Visual Content Package", "—", "Generic"),
    ("1", "L10", "Build Your First App", "My First App (Artifacts)", "—", "Generic"),
    ("1", "L11", "Build Something Bigger — Personal Utility", "Personal Utility (Lovable)", "cheatprompt.dev; tokenocd.com; goalsfor.me", "OK"),
    ("1", "L12", "Build Your First Website", "My Website (Lovable/v0)", "eyeonai.dev; Research+comparison template (14 domains); bestmtd.com; aicoursedirectory.com; Italian citizenship; partysnapper.com", "OK"),
    ("1", "L13", "Personal Finance Analysis", "Personal Finance analysis", "—", "No project — consider adding"),
    ("1", "L14", "Data Analysis — Dashboard", "My Dashboard", "tokenocd.com", "OK"),
    ("1", "L15", "Meet the Agents", "First Agent Task (Cowork)", "groceryshoppingagent.com; recipeunboxed.com", "OK"),
    ("1", "L16", "Custom GPTs & Claude Projects", "Personal AI Assistant", "bragmanager.com; goalsfor.me", "OK"),
    ("1", "L17", "Automation Basics", "Save Yourself an Hour (Zapier)", "groceryshoppingagent.com; recipeunboxed.com", "OK"),
    ("1", "L18-L21", "Family AI Bot Capstone", "Family AI Bot", "**familyaibot.com (Capstone)**", "OK"),
]

M2_LESSONS = [
    ("2", "L1", "Welcome — Vibe to Agentic", "Month 2 Charter doc", "—", "Meta / not a domain project"),
    ("2", "L2", "Professional Builder Stack", "Dev Env + First Commit", "—", "Setup"),
    ("2", "L3", "Context Engineering", "Re-spec a M1 App", "—", "Exercise"),
    ("2", "L4", "Calling Model APIs", "LLM API Playground & Cost Comparison", "**aiapi.click** (+ apicompare.net, llmapi.click, llmapicompare.com)", "OK — named project"),
    ("2", "L5", "Security — Keys, Injection, GDPR, ICO", "Hardened Project Template", "—", "Template, not public"),
    ("2", "L6", "Ingesting Real Business Data", "Document Ingestion Pipeline", "**Phase 1 of askmyco.com**", "OK — grown project"),
    ("2", "L7", "Vectors & Embeddings", "Semantic Search Demo", "—", "Optional demo"),
    ("2", "L8", "Build Your First RAG App", "Company Knowledge Bot", "askmyco.com (demoted from Capstone); aitranscriptionhub.com; cdiffguide.com; employmentlawbuddy.com", "OK"),
    ("2", "L9", "Making RAG Work (chunking + rerank)", "RAG Tuning Notebook", "(attaches to L8 projects)", "OK"),
    ("2", "L10", "Persistent Data + Multi-User", "Multi-User App Skeleton", "healthlog.app; jobtryout.co.uk", "OK"),
    ("2", "L11", "Agents — Claude SDK, CLIs, MCP", "Fractional Inbox Assistant", "**Phase 1 of fractionalbuddy.com**; traincoder.com", "OK — grown project"),
    ("2", "L12", "Browser & Computer-Use Agents", "Planning Applications Research Agent", "**planningapplicationbot.co.uk**; sharetrajectory.com; stagsguide/hensguide; newsbot.biz; blogbot.app", "OK"),
    ("2", "L13", "Voice Agents", "First Voice Agent (ElevenLabs)", "askevery.one (moved from M3 Capstone); Spanish GCSE Voice Tutor; jobtryout voice interview; myvideo.cv voice CV", "OK"),
    ("2", "L14", "Multimodal Pipelines", "Content Pipeline (product page)", "whiteboard.bot; talktosleep.com (Optional); myvideo.cv", "OK"),
    ("2", "L15", "Production Automation (n8n 2.0)", "Multi-Step Automation (n8n+Stripe+Supabase)", "**Phase 3 of sitegeo.net**; newsbot.biz; blogbot.app", "OK — grown project"),
    ("2", "L16-L19", "AI Readiness Assessment Tool (Capstone)", "Capstone (parts 1-4)", "**productarchitect.dev (NEW M2 Capstone)**", "REPLACED old askmyco capstone"),
    ("2", "L20", "Consulting Service Page", "Consulting Service Page", "fractionalbuddy.com", "OK"),
    ("2", "Lab", "M2-Lab 5 — Local STT", "Local Whisper / AnythingLLM", "lowhisper.com", "OK"),
]

M3_LESSONS = [
    ("3", "L1", "20/75 Rule", "(strategy)", "—", "—"),
    ("3", "L2", "12 Themes", "(strategy — feeds Capstone scoring)", "—", "Rolled into Capstone"),
    ("3", "L5", "Find Leverage Points", "Domain Discovery Assistant", "ideagrading.com", "OK"),
    ("3", "L6", "Domains not Use Cases", "Domain Discovery Assistant", "ideagrading.com", "OK"),
    ("3", "L10", "Harness Engineering (Ramp Glass)", "Internal Skills Library (Dojo)", "**aiconfident.io** (+ aiserviceindex.com, corporateaihub.com)", "OK — domain assigned"),
    ("3", "L14", "Proprietary Data Advantage", "(strategy)", "—", "—"),
    ("3", "L15", "Right to Deploy", "Right-to-Deploy Template Generator", "**governancekit.co.uk**", "OK — domain assigned"),
    ("3", "L1-L20", "AI Readiness Assessment Tool (Capstone)", "Extended Enterprise Edition", "**productarchitect.dev (same M2 capstone evolved)**", "REPLACED old askevery.one capstone"),
]

row_idx = 2
for row in M1_LESSONS + M2_LESSONS + M3_LESSONS:
    for col_idx, val in enumerate(row, start=1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = wrap
        c.border = border
        c.font = Font(size=10)
    month_fill_map = {
        '1': PatternFill("solid", fgColor="E8F8F0"),
        '2': PatternFill("solid", fgColor="E3F4FF"),
        '3': PatternFill("solid", fgColor="E8F7EE"),
    }
    fill = month_fill_map.get(row[0])
    if fill:
        for c_idx in range(1, len(ws2_headers) + 1):
            ws2.cell(row=row_idx, column=c_idx).fill = fill
    row_idx += 1

ws2_widths = [7, 10, 40, 45, 50, 35]
for i, w in enumerate(ws2_widths, start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws2.freeze_panes = "A2"


# ---------- SHEET 3: CHANGE LOG ----------
ws3 = wb.create_sheet("Change Log")
ws3_headers = ["#", "Change", "Why", "Files cross-referenced"]
for col_idx, h in enumerate(ws3_headers, start=1):
    c = ws3.cell(row=1, column=col_idx, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = border

CHANGES = [
    ("1", "**Replaced M2 Capstone**: old 'AI Customer-Support Chatbot / askmyco.com' -> new 'AI Readiness Assessment Tool / productarchitect.dev'. askmyco.com demoted to L8 Company Knowledge Bot.",
     "The April 2026 Month-2 redesign (gwth_lesson_ideas/MONTH_2_LESSON_IDEAS_2026-04-21.md) puts the Assessment Tool in M2 because its tech (Next.js + Supabase + RAG + agents + PDF + deploy) is exactly what M2 teaches. M3 then uses the same codebase to apply strategic frameworks.",
     "MONTH_2_LESSON_IDEAS_2026-04-21.md §Capstone; 20-capstone-architecture.md"),
    ("2", "**Replaced M3 Capstone**: old 'AI Readiness Voice Agent / askevery.one' -> new 'AI Readiness Assessment Tool — Enterprise edition' (evolution of M2 tool). askevery.one moved to M2 L13 as a voice-intake feature that feeds the Assessment Tool.",
     "New M3 lesson plan evolves the M2 capstone rather than building a new one. Voice agents are taught in M2 L13, so askevery.one fits naturally there.",
     "MONTH_3_LESSON_IDEAS_2026-04-20.md §Capstone; MONTH_2_LESSON_IDEAS_2026-04-21.md §L13"),
    ("3", "**Merged duplicates**: Row 15 cheatprompt.dev (M1) + Row 36 aicheatprompts.com (M2) were the same prompt-hub idea. Kept M1 (simpler Lovable build).",
     "Single source of truth; eliminates cross-month confusion.",
     "—"),
    ("4", "**Consolidated 14 comparison sites** (Rows 17, 33, 60-69) into ONE M1 'Research + Comparison Site Template' with the full domain list attached.",
     "Same build pattern, different niches - one row is cleaner than fourteen.",
     "MONTH_1_LESSON_IDEAS_2026-04-20.md §L12"),
    ("5", "**Month reassignments**: Row 13 blogbot.app M1->M2 (needs agents); Row 16 aitranscriptionhub.com M3->M2 (L8 RAG); Row 37 newsbot.biz M1->M2 (needs agents+automation); Row 10 lowhisper.com M1->M2 Lab (local STT).",
     "Each project's tech requirement places it in the correct month. Earlier allocations predated the curriculum redesign.",
     "MONTH_2_LESSON_IDEAS_2026-04-21.md Week 5-7"),
    ("6", "**Added regulatory risk flags**: MHRA (medical, cdiffguide), SRA / 'not legal advice' (employmentlawbuddy), FCA (sharetrajectory), GDPR special-category (healthlog), UK Employment Law + bias (jobtryout), IR35 / PI (fractionalbuddy), CAP Code affiliate disclosure (comparison template), copyright/fair-dealing (newsbot, blogbot).",
     "UK-specific compliance is a Month 2 lesson (L5) - project risk columns must reflect real regulatory exposure before students launch.",
     "MONTH_2_LESSON_IDEAS_2026-04-21.md §L5"),
    ("7", "**Added 7 missing projects** (from lesson ideas that weren't in the sheet): M2 L4 API-Powered Mini-App; M2 L6 Document Parser; M2 L11 MCP Email-Triage Agent; M2 L12 Planning Applications Research Agent; M2 L15 Production Automation Stack; M3 Mini-Build 3 Right-to-Deploy Generator; M3 Mini-Build 4 Internal AI Skills Library (Dojo).",
     "Ensures every build in the lesson plans has a corresponding project slot in the spreadsheet. Domains left as '(David to choose)' to avoid unilateral decisions.",
     "MONTH_2_LESSON_IDEAS_2026-04-21.md §Build projects; MONTH_3_LESSON_IDEAS_2026-04-20.md §Mini-Builds"),
    ("8", "**Added columns**: Status, Lesson(s), Primary Tools, Alternative Tools, Monetisation, Separate Site?, Complexity, Dependencies, Change Notes.",
     "Original sheet had only 9 columns, most blank. New structure captures monetisation (Free vs Paid - matters because some sites need to cover AI token costs), separate-site status (advertising for GWTH), and dependency chain (lesson X taught before project Y).",
     "User request"),
    ("9", "**Retained David's prior learnings in Risks column**: newsbot.biz - 'voting feature tested, zero engagement, drop it'; fractionalbuddy - prototype built, customers interested.",
     "Preserves institutional memory.",
     "Original sheet rows 35, 37"),
    ("10", "**Marked Optional/Lab**: talktosleep.com reclassified from M2 Core to M2 Optional/Lab because Suno is not in the prescriptive big-3 stack.",
     "Prescriptive-stack rule from MONTH_2_LESSON_IDEAS §Appendix C - unproven tools go to Labs, not Core lessons.",
     "MONTH_2_LESSON_IDEAS_2026-04-21.md §tooling rules"),
    ("11", "**v4 update — David's 21/4 feedback pass**: 6 direct domain assignments (aicoursedirectory.com rebuild; spanish.gcseapp.com; traincoder.com; planningapplicationbot.co.uk; aiconfident.io for Dojo; governancekit.co.uk for Right-to-Deploy).",
     "Applied David's feedback verbatim; used his 488-domain list as source of truth.",
     "Domain_List (1).csv; v3 annotated spreadsheet"),
    ("12", "**v4 update — reframed 4 'too generic' builds** per David's 21/4 feedback: (L4) API Mini-App -> LLM API Playground on aiapi.click; (L6) Document Parser -> Phase 1 of askmyco.com; (L11) MCP Email-Triage -> Phase 1 of fractionalbuddy.com (Fractional Inbox Assistant); (L15) Production Automation -> Phase 3 of sitegeo.net.",
     "David's note: 'Too generic - pick a project from above that this can be integrated with and maybe call this phase 1 or phase 2'. Result: every lesson now teaches against a named, domain-owned project — no abstract demos.",
     "Domain_List (1).csv; v3 annotated spreadsheet"),
]

row_idx = 2
for row in CHANGES:
    for col_idx, val in enumerate(row, start=1):
        c = ws3.cell(row=row_idx, column=col_idx, value=val)
        c.alignment = wrap
        c.border = border
        c.font = Font(size=10)
    row_idx += 1

ws3_widths = [5, 80, 70, 50]
for i, w in enumerate(ws3_widths, start=1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws3.freeze_panes = "A2"

out = r"C:\Users\david\Downloads\All GWTH Project Ideas 21_Apr_26 v4.xlsx"
wb.save(out)
print(f"Saved: {out}")
